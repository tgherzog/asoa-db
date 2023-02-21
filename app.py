
import flask
from werkzeug.middleware.proxy_fix import ProxyFix
from flask_assets import Environment, Bundle
import markupsafe
from datetime import datetime
from markdown import markdown
import os
import io
import tempfile

import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

config = {
  'db_path': 'data/asoa-roster.xlsx'
}

config['db_lastmod'] = xl.load_workbook(filename=config['db_path']).properties.modified

# set "asoa_access_mode=members" as an environment variable to enable privileged access
# to member-only database content and features
config['access_mode'] = os.environ.get('asoa_access_mode', 'public')

app = flask.Flask(__name__)
# if being run via gunicorn, make us proxy-aware (assume Apache)
if os.environ.get('_', '').endswith('gunicorn'):
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_prefix=1)
assets = Environment(app)
# assets.url = app.static_url_path
assets.append_path('assets/scss')
assets.url_expire = False   # disable dynamic CSS loading
scss = Bundle('custom.scss', filters='pyscss', output='css/custom.css')
assets.register('scss_site', scss)


@app.template_filter('markdown')
def app_markdown_filter(s):
    
    return markupsafe.Markup(markdown(s))

@app.context_processor
def context_processor():
    request = flask.request

    email = 'seabreezeowners@gmail.com'
    public_url = request.url.split('://')[1]
    if request.host.split('.')[0] == 'asoadb':
        private_url = public_url.split('.', maxsplit=1)
        private_url[0] = 'members'
        private_url = '.'.join(private_url)
    else:
        private_url = public_url

    def email_link(title=None):
        if not title:
            title = email

        return markupsafe.Markup('<a href="mailto:{}">{}</a>'.format(email, title))

    return {
      # global variables
      'access_mode': config['access_mode'],
      'asoa_email': email,
      'public_url': public_url,
      'private_url': private_url,
      'db_datestamp': config['db_lastmod'].strftime('%m/%d/%Y %I:%M%p UTC'),
      'db_update_form': 'https://forms.gle/usB89vY9sc5U6adG8',

      # custom functions
      'contact_asoa': email_link
    }


@app.route('/')
def app_start():

  return flask.render_template('start.html')

@app.route('/list')
def app_list():

    return flask.render_template('list.html', page_title='List of All Known Seabreezes', boats=db_load())

@app.route('/detail/<hull>')
def app_detail(hull):

    boat = db_load(hull)
    if boat:
        # we also pass the original hull number in case of error
        return flask.render_template('detail.html', hull=hull, boat=db_load(hull))

    flask.abort(404, 'Hull {} is not a known Seabreeze'.format(hull))

@app.route('/search')
def app_search():
    
    q = flask.request.args.get('q')
    if q:
        try:
            q = int(q)
        except:
            pass

        if type(q) is int:
            return flask.redirect('/detail/{}'.format(q))
        else:
            return flask.render_template('list.html', search_term=q, page_title='Search Results', boats=db_load(q=q))

    return flask.redirect(flask.url_for('app_list'))

def download_member_file():

    wb = xl.Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = 'members'
    hdrFont = Font(bold=True)

    fields = ['hull', 'boat_name', 'status', 'sailnum', 'rig', 'owner_name', 'acquired', 'address1', 'address2', 'phone', 'email', 'berth', 'latest_info', 'epitaph']
    for n in range(len(fields)):
        cell = ws.cell(1, n+1)
        cell.value = fields[n]
        cell.font = hdrFont
        ws.column_dimensions[get_column_letter(n+1)].width = 10

    ws.cell(1, fields.index('hull')+1).alignment = Alignment(horizontal='center')

    row = 2

    for hull,boat in db_load(raw=True).items():
        owner = boat['owners'][0] if len(boat['owners']) > 0 else {}
        for k in ['owner_name', 'acquired', 'address1', 'address2', 'phone', 'email']:
            boat[k] = owner.get(k)

        for n in range(len(fields)):
            ws.cell(row, n+1).value = boat[fields[n]]

        # improve formatting of a few  columns
        ws.cell(row, fields.index('acquired')+1).number_format = 'm/d/yyyy;@'
        ws.cell(row, fields.index('hull')+1).alignment = Alignment(horizontal='center')
        row += 1

    # spiff up the column formatting
    ws.column_dimensions[get_column_letter(fields.index('boat_name')+1)].width *= 2
    ws.column_dimensions[get_column_letter(fields.index('owner_name')+1)].width *= 2.5
    ws.column_dimensions[get_column_letter(fields.index('address1')+1)].width *= 2
    ws.column_dimensions[get_column_letter(fields.index('address2')+1)].width *= 2
    ws.column_dimensions[get_column_letter(fields.index('email')+1)].width *= 2
    ws.column_dimensions[get_column_letter(fields.index('berth')+1)].width *= 2
    ws.column_dimensions[get_column_letter(fields.index('phone')+1)].width *= 1.5
    ws.column_dimensions[get_column_letter(fields.index('acquired')+1)].width *= 1.25

    ws.freeze_panes = ws['A2']

    with tempfile.NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        tmp.close()
        byteStream = io.BytesIO(stream)

        filename = 'asoa-members-{}.xlsx'.format(config['db_lastmod'].strftime('%Y-%m-%d'))
        return flask.send_file(byteStream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=filename, as_attachment=True)


if config['access_mode'] == 'members':
    app.add_url_rule('/download/members', view_func=download_member_file)


def db_load(id=None, q=None, raw=False):
    '''Loads the database from a spreadsheet.
      
       id:      hull number. If defined, the function returns a single record (dict), otherwise
                a list of records (dicts)

       q:       search text to filter results. Else return all boats

    '''

    def filter_boats(item):
        '''Crude search functionality: search on boat name, berth, and owners
        '''

        (k,boat) = item
        s = ' '.join([boat['boat_name'], boat['berth']] + list(map(lambda x: x['owner_name'], boat['owners'])))
        return q in s.lower()

    def fmt_date(date):
        if raw:
            return date

        if type(date) is datetime:
            # 7/1 is a special date signifying that only the year has any confidence
            # mm/1 means that only month and year have confidence so we omit the date
            if date.month == 7 and date.day == 1:
                return date.strftime('%Y')
            elif date.day == 1:
                return date.strftime('%-m/%Y')

            return date.strftime('%-m/%-d/%Y')

        return ''    

    wb = xl.load_workbook(filename=config['db_path'], read_only=True, data_only=True)
    boats = wb['boats']
    boat_db = {}
    keys = {}
    i = 0
    for col in boats[1]:
        keys[col.value] = i
        i += 1

    for row in boats.iter_rows(2):
        boat = {}
        for key in ['hull', 'date', 'status', 'boat_name', 'sale_link', 'sailnum', 'rig', 'color', 'engine_type', 'engine_desc', 'berth', 'epitaph', 'latest_info']:
            boat[key] = row[keys[key]].value or ''

        if not raw:
            boat['hull'] = str(boat['hull'])

        if not boat['hull']:
            continue

        owner = {'hull': boat['hull'], 'acquired': fmt_date(boat['date'])}

        for key in ['owner_name', 'address1', 'address2', 'phone', 'email']:
            owner[key] = row[keys[key]].value or ''

        boat['date'] = fmt_date(boat['date'])

        hull = boat['hull']
        if boat_db.get(hull):
            # Prior record exists: append owner to front and merge more recent fields
            if boat['status'] in ['GOOD', 'RENO']:    
                boat_db[hull]['owners'].insert(0, owner)

            for k,v in boat.items():
                if v:
                    boat_db[hull][k] = v
        else:
            boat['owners'] = []
            if boat['status'] in ['GOOD', 'RENO']:
                boat['owners'].append(owner)

            boat_db[hull] = boat

    if id:
        return boat_db.get(id, {})
    elif q:
        q = q.lower()
        return dict(filter(filter_boats, boat_db.items()))

    return boat_db
